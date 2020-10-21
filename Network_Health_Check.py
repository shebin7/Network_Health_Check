from nornir import InitNornir
from nornir_netmiko.tasks import netmiko_send_command,netmiko_send_config
from nornir_utils.plugins.functions import print_result,print_title
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.prompt import Prompt,IntPrompt
from netmiko import ConnectHandler
from openpyxl import load_workbook
from openpyxl.styles import *
from alive_progress import alive_bar
from datetime import *
import textfsm
import openpyxl
import time
import csv


### Displays System Time when called ###
date_today = datetime.today()
time_now = datetime.now()
Dtime = time_now.strftime("%H:%M:%S")


### File path for config.yaml ###
config_file_yaml="/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/config.yaml"

### File path for all the Branch Ip Address which we will ping from central server to check connectivity ###
branch_ip_address_for_pinging = '/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/branch_ipaddress_2.csv'

### Final result Device Up or Down result will be stored here  in csv ###
final_result_of_Health_Check="/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/"+str(date_today)+'__'+"final_result.csv"

#### Final Result Path of Excell file with color based on Device Status ###
ping_result="/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/"+str(date_today)+'__'+"ping_result.xlsx"


### Initilizing nornir object ###
nr = InitNornir(config_file="/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/config.yaml")


### Reading Total number of Rows in CSV files ###
file = open(branch_ip_address_for_pinging)
num_rows = len(file.readlines())


### Rich Variable for styles&tabels ###
status_up  = '[bold] :white_heavy_check_mark: '
status_down = '[bold] :cross_mark:'

style_down = '[red][bold] '
style_up = '[green] [bold] '


### Creating a Table for the result on CLI ###
console = Console()
cli_table = Table(title='[bold]Network Health Check Table [/bold]')
cli_table.add_column("SOL_ID", justify="right", style="cyan", no_wrap=True)
cli_table.add_column("BRANCH_NAME", justify="right", style="cyan", no_wrap=True)
cli_table.add_column("IP_ADDRESS", justify="right", style="cyan", no_wrap=True)
cli_table.add_column("HEALTH_STATUS", justify="right", style="cyan", no_wrap=True)


### Header fields for csv file where ping result will be stored ###
report_fields = ['SOL_ID','IP_ADDRESS','BRANCH_NAME','HEALTH_STATUS','TIME']


### Opening and writing Headers to Csv File ###
with open(final_result_of_Health_Check,'a')as wr:
    csv_dictwrite = csv.DictWriter(wr,report_fields)
    csv_dictwrite.writeheader()



console.print(Panel("[bold]  :smiley:   ~WELCOME TO NETWORK HEALTH CHECKUP~   :smiley: ",style='white on blue'),justify='center')
console.print()
console.print()
#console.print(':computer: [green][bold] `~`~`~`~`~`~`~`~`~`~`~`~`[/bold][/green]:satellite: ~`~`~`~`~`~`~`~`~`~`~`~` :computer:   ',justify='center')  
console.print()
console.print(' :satellite:','[blink]:sparkles:'*4,                       ':satellite:',justify='center')
console.print(' :vertical_traffic_light:'+'          '+':vertical_traffic_light:',' '*14,justify='center')   
console.print(':bank:',':traffic_light:'*10,':cloud:'+'           '+':cloud:',':traffic_light:'*10,':office:',justify='center' )  
console.print()
console.print()
console.print()
console.print("[bold][red][blink]Warning![/blink][/red][/bold] [red][italic]PLEASE DONOT INTERUPT THE EXECUTION",justify='center')
console.print()
console.print()
console.print('Working...',style='green')
time.sleep(0.5)


def Intermediate_Server_ping_pattern_catcher(result_string,device_status):
    device_status = '' 
    template_path ='/home/shebin/NETDEVOPS/Net_automation_Project/Network_Health_Check/ping_linux.textfsm'
    with open(template_path,'r')as tm:
        ping_template = textfsm.TextFSM(tm)

    result_ping  = ping_template.ParseText(result_string)

    template_header= ping_template.header

    dict_result = [dict(zip(template_header,i))for i in result_ping]

    count_exceed=0
    count_unreach=0
    count_ok=0  
    for i in dict_result:
        if 'time' in str(i['result']):
            count_ok=count_ok+1
        elif 'Destination' in str(i['result']):
            count_unreach =count_unreach+1
        elif 'Time' in str(i['result']):
            count_exceed = count_exceed+1

    if count_ok > 6:
        return device_status+'UP'
        
    elif (count_unreach > 8) or (count_exceed > 8):
        return device_status+'DOWN'
    


prompt_str='\n IF YOU DONT KNOW THE HOSTNAME AND PORT TO [bold][blue]<CONNECT DEVICE>[/bold][/blue] JUST PRESS [bold][blue]<ENTER> [/bold][/blue]DEFAULT VALUES WILL BE SELECTED \n'
console.print('Do you have Access to Network Devices through',style='bold purple')
console.print('\n[white]:right_arrow:[/white]JUMP/INTERMEDIATE SERVER ACCESS',style='bold green')
console.print('[white]:right_arrow:[/white]DIRECT SYSTEM ACCESS\n',style='bold green')
Select_Method = Prompt.ask("[bold yellow]SELECT JUMP/INTERMEDIATE SERVER [bold purple]OR[/bold purple DIRECT ACESS=",choices=['Intermediate','Direct'])




prompt_str='\n IF YOU DONT KNOW THE HOSTNAME AND PORT TO [bold][blue]<CONNECT DEVICE>[/bold][/blue] JUST PRESS [bold][blue]<ENTER> [/bold][/blue]DEFAULT VALUES WILL BE SELECTED \n'
server_ip=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server IP ADDRESS=',default='192.168.4.133')
server_username=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server USERNAME=',default='shebin')
server_password=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server IP PASSWORD=',default='shebin123',show_default=False)

if 'Intermediate' in Select_Method:
    def Intermediate_Server():

        global server_ip,server_password,server_username
        #prompt_str='\n IF YOU DONT KNOW THE HOSTNAME AND PORT TO [bold][blue]<CONNECT DEVICE>[/bold][/blue] JUST PRESS [bold][blue]<ENTER> [/bold][/blue]DEFAULT VALUES WILL BE SELECTED \n'
        #server_ip=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server IP ADDRESS=',default='192.168.4.133')
        #server_username=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server USERNAME=',default='shebin')
        #server_password=Prompt.ask('\n [bold yellow]Type your Intermediate/Jump Server IP PASSWORD=',default='shebin123',show_default=False)

        jump_server={'device_type':'terminal_server','ip':str(server_ip),'username':str(server_username),
            'password':str(server_password),'global_delay_factor':1}
        net_connect = ConnectHandler(**jump_server)
        print(net_connect.is_alive())

        try:
            net_connect= ConnectHandler(**jump_server)
            with alive_bar(num_rows-1)as bar:
                with open(branch_ip_address_for_pinging,'r')as b_ip:
                    csv_d_reader = csv.DictReader(b_ip)
                    for row in csv_d_reader:
                        row_values={'ip' : row['IP_ADDRESS'],'solid' : row['Sol_ID'],'branch' : row['Branch_Name']}
           
                        net_connect.write_channel('ping '+row['IP_ADDRESS']+' -c 10 \n')

                        write_channel_op_1=net_connect._read_channel_timing(delay_factor=2,max_loops=150)
                        print(write_channel_op_1)  
 
                        Device_Status=Intermediate_Server_ping_pattern_catcher(result_string=write_channel_op_1,device_status='')
                        Device_Status
                        print(Device_Status)

                        if Device_Status == 'DOWN':

                            with open(final_result_of_Health_Check,'a+')as wr:

                                csv_dictwrite = csv.DictWriter(wr,report_fields)                        
                                csv_dictwrite.writerow({'SOL_ID':row['Sol_ID'],'IP_ADDRESS': row['IP_ADDRESS'],'BRANCH_NAME':row['Branch_Name'],'HEALTH_STATUS':'DOWN','TIME':Dtime})
                                sol_id_down = style_down+row_values['solid']
                                branch_down = style_down+row_values['branch']
                                ip_down = style_down+row_values['ip']
                                cli_table.add_row(sol_id_down,branch_down,ip_down,status_down)
                
                        else:
                    
                            with open(final_result_of_Health_Check,'a+')as wr:

                                csv_dictwrite = csv.DictWriter(wr,report_fields)                        
                                csv_dictwrite.writerow({'SOL_ID':row['Sol_ID'],'IP_ADDRESS': row['IP_ADDRESS'],'BRANCH_NAME':row['Branch_Name'],'HEALTH_STATUS':'DOWN','TIME':Dtime})
                                sol_id_up = style_up+row_values['solid']
                                branch_up = style_up+row_values['branch']
                                ip_up = style_down+row_values['ip']
                                cli_table.add_row(sol_id_up,branch_up,ip_up,status_up)


### Bar tracks/shows  ETA/Time needed to complete the execution ###   
 
                        bar()
                        continue
            

        except ConnectionError:
            print('Exception')

        finally:
            net_connect.disconnect()
 
    Intermediate_Server() 


elif 'Direct' in Select_Method:
    def Device_as_Server(task):
        with alive_bar(num_rows-1)as bar:
            with open(branch_ip_address_for_pinging,'r')as re:
                csv_d_reader = csv.DictReader(re)
                for row in csv_d_reader:
                 
                    row_values={'ip' : row['IP_ADDRESS'],'solid' : row['Sol_ID'],'branch' : row['Branch_Name']}
                    res = task.run(netmiko_send_command,command_string='ping '+row_values['ip'])
                    ping_search = res.result
                    if not '!!!' in ping_search:

                        ### Opening New CSV and writing individual device 'Status' and their details ###
                        with open(final_result_of_Health_Check,'a+')as wr:
                            csv_dictwrite = csv.DictWriter(wr,report_fields)
                            csv_dictwrite.writerow({'SOL_ID':row['Sol_ID'],'IP_ADDRESS': row['IP_ADDRESS'],'BRANCH_NAME':row['Branch_Name'],'HEALTH_STATUS':'DOWN','TIME':Dtime})
                                        
                            sol_id_down = style_down+row_values['solid']
                            branch_down = style_down+row_values['branch']
                            ip_down = style_down+row_values['ip']
                            cli_table.add_row(sol_id_down,branch_down,ip_down,status_down)
                                                      
                    else:                    
                        ### Opening New CSV and writing individual device 'Status' and their details ###
                        with open(final_result_of_Health_Check,'a+')as wr:
                            csv_dictwrite = csv.DictWriter(wr,report_fields)
                            csv_dictwrite.writerow({'SOL_ID':row['Sol_ID'],'IP_ADDRESS': row['IP_ADDRESS'],'BRANCH_NAME':row['Branch_Name'],'HEALTH_STATUS':'UP','TIME':Dtime})

                    
                            sol_id_up = style_up+row_values['solid']
                            branch_up = style_up+row_values['branch']
                            ip_up = style_up+row_values['ip']
                            cli_table.add_row(sol_id_up,branch_up,ip_up,status_up)



### Bar tracks/shows  ETA/Time needed to complete the execution ###    
                    bar()
         
                    continue


    result_of_ping = nr.run(Device_as_Server)        
    result_of_ping


console.print(cli_table)

### Now writing csv rows to Excell sheet and filling it with colors based on Status ###

wb = openpyxl.Workbook()
excell_sheet = wb.active


with open(final_result_of_Health_Check,'r')as f:
    reader = csv.reader(f)
    for row in reader:
        excell_sheet.append(row)

wb.save(filename=ping_result)


wk = load_workbook(ping_result)
ws = wk.active

### Counting maximum rows and coloums of excell_sheet ###

max_rows = ws.max_row
max_cols = ws.max_column

### Iterating over excell rows and coloums to find Status  and fill it with color ###
for r in range(1,max_rows+1):
    for c in range(1,max_cols):
        c_val = ws.cell(row=r,column=4).value
        if c_val ==('DOWN' or  'down' or 'Down'):
            ws.cell(row=r,column=4).fill=PatternFill(fgColor='FF0000',fill_type='solid')
        elif c_val==('UP' or  'up' or 'Up'):
            ws.cell(row=r,column=4).fill=PatternFill(fgColor='00FF00',fill_type='solid')
        else:
            pass

### Saving the Changes ###
wk.save(ping_result)

